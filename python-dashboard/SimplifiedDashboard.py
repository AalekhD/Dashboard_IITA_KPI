import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from openpyxl.utils import get_column_letter
import os
import numpy as np

# Page config
st.set_page_config(page_title="IITA KPI Dashboard", layout="wide")

# Header
st.markdown("""
<div style="background-color:#00891a; padding:20px; border-radius:10px;">
    <h1 style="color:#ffffff; text-align:center; margin:0;">🌱 IITA KPI Dashboard</h1>
    <p style="color:#ffffff; text-align:center; margin:5px;">IITA Programs and Service Unit KPIs</p>
</div>
""", unsafe_allow_html=True)

st.write("")

# Load Excel files and convert to HTML with merged cells
@st.cache_data
def load_kpi_data():
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    # Load Program Output KPIs
    program_file = os.path.join(root_dir, 'data', 'Program Output KPIs.xlsx')
    df_programs = pd.read_excel(program_file)
    
    # Load Service Unit KPIs
    service_file = os.path.join(root_dir, 'data', 'Service Unit KPIs.xlsx')
    df_services = pd.read_excel(service_file)
    
    # Load KPI Heat map
    heatmap_file = os.path.join(root_dir, 'data', 'KPI by Nr. Heat map.xlsx')
    df_heatmap = pd.read_excel(heatmap_file)
    
    return df_programs, df_services, df_heatmap

# Function to convert Excel with merged cells to HTML
def excel_to_html_with_merged_cells(excel_file_path, no_decimals=False, highlight_row_keyword=None):
    # Load workbook with data_only=True to get calculated values instead of formulas
    wb_data = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws_data = wb_data.active
    
    # Load workbook normally to get formatting info (merged cells, number formats)
    wb_format = openpyxl.load_workbook(excel_file_path)
    ws_format = wb_format.active
    
    # Find actual data range (trim empty rows and columns)
    max_row = 0
    max_col = 0
    
    for row_idx, row in enumerate(ws_data.iter_rows(values_only=True), 1):
        has_data = any(cell is not None for cell in row)
        if has_data:
            max_row = row_idx
            # Find max column with data in this row
            for col_idx, cell in enumerate(row, 1):
                if cell is not None:
                    max_col = max(max_col, col_idx)
    
    # Get merged cell ranges from format workbook
    merged_cells = {}
    for merged_range in ws_format.merged_cells.ranges:
        cells = list(merged_range.cells)
        for cell in cells[1:]:  # Skip the first cell (top-left)
            merged_cells[cell] = cells[0]  # Map to top-left cell
    
    # Build HTML table
    html = '<table style="border-collapse: collapse; width: 100%; table-layout: fixed; font-family: Arial, sans-serif; background-color: white;">'
    html += '<style>td, th { border: 1px solid #999; padding: 12px; text-align: left; background-color: white; white-space: normal; word-wrap: break-word; word-break: break-word; overflow-wrap: break-word; color: black; font-size: 9pt; }</style>'
    
    processed = set()
    
    for row_idx in range(1, max_row + 1):
        # Get the actual row
        row_data = list(ws_data.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
        
        # Check if this row is completely empty
        has_row_data = False
        for col_idx in range(1, max_col + 1):
            if row_data[col_idx - 1].value is not None:
                has_row_data = True
                break
        
        # Skip completely empty rows
        if not has_row_data:
            continue

        # Determine if this row should be highlighted (contains keyword)
        highlight_row = False
        if highlight_row_keyword:
            try:
                lowkw = highlight_row_keyword.strip().lower()
                for c in row_data:
                    if c.value is not None and lowkw in str(c.value).lower():
                        highlight_row = True
                        break
            except Exception:
                highlight_row = False

        html += '<tr>'
        for col_idx in range(1, max_col + 1):
            cell_data = row_data[col_idx - 1]
            cell_coord = cell_data.coordinate
            
            # Get corresponding format cell
            cell_format = ws_format[cell_coord]
            
            # Skip if this cell is part of a merged range (not the top-left)
            if cell_coord in merged_cells and merged_cells[cell_coord] != cell_coord:
                continue
            
            # Skip if already processed
            if cell_coord in processed:
                continue
            
            # Calculate rowspan and colspan for merged cells
            rowspan = 1
            colspan = 1
            
            for merged_range in ws_format.merged_cells.ranges:
                if cell_coord in merged_range:
                    rowspan = merged_range.max_row - merged_range.min_row + 1
                    colspan = merged_range.max_col - merged_range.min_col + 1
                    # Mark all cells in this range as processed
                    for r in range(merged_range.min_row, merged_range.max_row + 1):
                        for c in range(merged_range.min_col, merged_range.max_col + 1):
                            processed.add(f"{get_column_letter(c)}{r}")
                    break
            
            # Get cell value from data workbook (contains calculated values, not formulas)
            cell_value = cell_data.value
            
            # Format based on cell number format
            if cell_value is not None:
                if isinstance(cell_value, (int, float)):
                    # Check if the cell has percentage format
                    if cell_format.number_format and '%' in cell_format.number_format:
                        if no_decimals:
                            cell_value = f"{cell_value * 100:.0f}%"
                        else:
                            cell_value = f"{cell_value * 100:.2f}%"
                    else:
                        if no_decimals:
                            # Round to nearest integer and show without decimals
                            try:
                                cell_value = f"{int(round(cell_value)):,}"
                            except Exception:
                                cell_value = str(cell_value)
                        else:
                            # Up to 3 decimal places; strip trailing zeros; add thousand commas
                            try:
                                if cell_value == int(cell_value):
                                    cell_value = f"{int(cell_value):,}"
                                else:
                                    cell_value = f"{cell_value:,.3f}".rstrip('0').rstrip('.')
                            except Exception:
                                cell_value = str(cell_value)
                else:
                    cell_value = str(cell_value)
            else:
                cell_value = ""
            
            # Add styling for headers (first row)
            if row_idx == 1:
                html += f'<th style="background-color: #00891a; color: white; font-weight: bold;" rowspan="{rowspan}" colspan="{colspan}">{cell_value}</th>'
            else:
                # If this row matches the highlight keyword, make its cells green with white text
                if highlight_row:
                    html += f'<td style="background-color: #00891a; color: white;" rowspan="{rowspan}" colspan="{colspan}">{cell_value}</td>'
                else:
                    html += f'<td rowspan="{rowspan}" colspan="{colspan}">{cell_value}</td>'
        
        html += '</tr>'
    
    html += '</table>'
    return html

# Function to create heatmap from KPI heat map file
# Returns (fig, df_below) where df_below contains rows beyond row 16 (or None)
def create_heatmap_visualization(excel_file_path, heatmap_max_row=16,
                                  data_col_start=4, program_col=3, group_col=2,
                                  kpi_row=3, kpi_group_row=2, data_row_start=4,
                                  show_row_groups=True, include_below_rows=True,
                                  left_margin=None, side_cols=None,
                                  zero_decimal_cols=None, one_decimal_cols=None, two_decimal_cols=None,
                                  zero_decimal_rows=None, one_decimal_rows=None, force_decimals=None, suppress_pct_display=False, monospace_numeric=False,
                                  one_decimal_first_col=False, no_gray_first_col=False,
                                  kpi_group_filter=None, force_include_cols=None, kpi_group_source=None, extra_top=0, group_gap=None):
    try:
        # Load with openpyxl to get clean numeric data
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        ws = wb.active

        # Build a map of merged cells -> top-left cell coordinates so we can
        # read values/number formats from the merged-region anchor when a
        # cell belongs to a merged range.
        merged_map = {}
        try:
            for mr in ws.merged_cells.ranges:
                min_r, min_c = mr.min_row, mr.min_col
                for rr in range(mr.min_row, mr.max_row + 1):
                    for cc in range(mr.min_col, mr.max_col + 1):
                        merged_map[(rr, cc)] = (min_r, min_c)
        except Exception:
            merged_map = {}

        def merged_cell_coord(r, c):
            return merged_map.get((r, c), (r, c))

        def merged_val(r, c):
            tr, tc = merged_cell_coord(r, c)
            return ws.cell(row=tr, column=tc).value

        def merged_cell_obj(r, c):
            tr, tc = merged_cell_coord(r, c)
            return ws.cell(row=tr, column=tc)

        programs = []
        program_groups = []   # column B: GI / RAFS / ST etc.
        kpi_names = []
        kpi_type_groups = []  # row 2: Research Outputs / Training etc.
        data_values = []
        original_values = []

        # Programs/rows below heatmap_max_row
        below_programs = []
        below_data = []
        below_orig = []

        # Build program-group map from group_col (fill-forward for merged cells)
        current_b = None
        b_values = {}
        if group_col is not None:
            for row_idx in range(data_row_start, ws.max_row + 1):
                val = merged_val(row_idx, group_col)
                if val is not None:
                    current_b = str(val)
                b_values[row_idx] = current_b

        # Get base KPI names and columns from kpi_row (starting from data_col_start)
        # Scan to last non-None header to avoid empty trailing columns. Use
        # merged_val so headers that are part of merged cells are detected.
        base_kpi_cols = []
        base_kpi_names = []
        last_kpi_col = data_col_start
        for col_idx in range(data_col_start, ws.max_column + 1):
            cell_val = merged_val(kpi_row, col_idx) if kpi_row else None
            if cell_val is not None:
                last_kpi_col = col_idx
        for col_idx in range(data_col_start, last_kpi_col + 1):
            if (side_cols or []) and col_idx in (side_cols or []):
                continue  # already handled as a side column — do not add to base KPI list
            cell_val = merged_val(kpi_row, col_idx) if kpi_row else None
            if cell_val is not None:
                base_kpi_cols.append(col_idx)
                base_kpi_names.append(str(cell_val))

        # side_cols are absolute Excel column indices to include as uncolored side columns
        side_cols = side_cols or []

        # Build final included columns and kpi_names: side_cols first, then base KPI cols
        included_cols = []
        kpi_names = []
        for c in side_cols:
            if kpi_row:
                obj = merged_cell_obj(kpi_row, c)
                hdr = obj.value if obj and obj.value is not None else get_column_letter(c)
            else:
                hdr = get_column_letter(c)
            included_cols.append(c)
            kpi_names.append(str(hdr))
        included_cols.extend(base_kpi_cols)
        kpi_names.extend(base_kpi_names)

        # Allow caller to force include specific absolute column indices
        # (useful for wide heatmap variants where data sits in widely spaced columns)
        if force_include_cols:
            for col_idx in force_include_cols:
                if col_idx not in included_cols and col_idx <= ws.max_column:
                    hdr = merged_val(kpi_row, col_idx) if kpi_row else None
                    included_cols.append(col_idx)
                    kpi_names.append(str(hdr) if hdr is not None else get_column_letter(col_idx))

        # Ensure we include any columns that contain data in the data rows even if the
        # KPI header cell is blank (some wide sheets leave header cells empty).
        for col_idx in range(data_col_start, ws.max_column + 1):
            if col_idx in included_cols:
                continue
            has_data = False
            # Check all data rows (including rows beyond heatmap_max_row) for any
            # non-empty cell in this column so we don't omit columns that only
            # have values in the 'below' region.
            for r in range(data_row_start, ws.max_row + 1):
                if merged_val(r, col_idx) is not None:
                    has_data = True
                    break
            if has_data:
                included_cols.append(col_idx)
                hdr = merged_val(kpi_row, col_idx) if kpi_row else None
                kpi_names.append(str(hdr) if hdr is not None else get_column_letter(col_idx))

        # Get KPI type groups from kpi_group_row (fill-forward for merged cells)
        current_type = None
        for c in included_cols:
            if c in side_cols:
                kpi_type_groups.append("")
            else:
                val = merged_val(kpi_group_row, c) if kpi_group_row else None
                if val is not None:
                    current_type = str(val)
                kpi_type_groups.append(current_type or "")

        # If a separate KPI group source file was provided, prefer its grouping
        # text (useful when Capacity/Product sheets omit the top grouping row).
        if kpi_group_source and os.path.exists(kpi_group_source):
            try:
                wb_src = openpyxl.load_workbook(kpi_group_source, data_only=True)
                ws_src = wb_src.active
                # Build merged map for source
                merged_map_src = {}
                try:
                    for mr in ws_src.merged_cells.ranges:
                        min_r, min_c = mr.min_row, mr.min_col
                        for rr in range(mr.min_row, mr.max_row + 1):
                            for cc in range(mr.min_col, mr.max_col + 1):
                                merged_map_src[(rr, cc)] = (min_r, min_c)
                except Exception:
                    merged_map_src = {}

                def merged_src_val(r, c):
                    tr, tc = merged_map_src.get((r, c), (r, c))
                    return ws_src.cell(row=tr, column=tc).value

                # Build group values from the source file with fill-forward semantics
                src_current = None
                src_groups = []
                for c in included_cols:
                    if c in side_cols:
                        src_groups.append("")
                    else:
                        v = merged_src_val(kpi_group_row, c) if kpi_group_row else None
                        if v is not None:
                            src_current = str(v)
                        src_groups.append(src_current or "")

                # If the source provided any non-empty group names, replace the groups
                if any(g for g in src_groups if g):
                    kpi_type_groups = src_groups
            except Exception:
                pass

        # If a KPI group filter was specified, filter included columns and names
        # by case-insensitive substring match on the KPI type group. This allows
        # creating focused "KPI over Time" views for Research Outputs,
        # Capacity Building, Product Development, etc.
        if kpi_group_filter:
            filt = str(kpi_group_filter).strip().lower()
            new_included = []
            new_names = []
            new_type_groups = []
            for col_idx, name, grp in zip(included_cols, kpi_names, kpi_type_groups):
                if grp and filt in grp.lower():
                    new_included.append(col_idx)
                    new_names.append(name)
                    new_type_groups.append(grp)
            if new_included:
                included_cols = new_included
                kpi_names = new_names
                kpi_type_groups = new_type_groups

        # Truncate to ~34 chars to match Excel row-3 height (191px) at 9pt font (~5.5px/char)
        def truncate_label(text, max_len=34):
            return text[:max_len] + '\u2026' if len(text) > max_len else text

        # Wrap label text at word boundaries, inserting <br> for Plotly HTML rendering
        def wrap_label(text, max_len=15):
            words = text.split()
            lines = []
            current = ''
            for word in words:
                if current and len(current) + 1 + len(word) > max_len:
                    lines.append(current)
                    current = word
                else:
                    current = (current + ' ' + word).strip()
            if current:
                lines.append(current)
            return '<br>'.join(lines)

        # Compute KPI type group spans for header annotations
        kpi_type_spans = []
        if kpi_type_groups:
            span_start = 0
            span_name = kpi_type_groups[0]
            for i, grp in enumerate(kpi_type_groups[1:], 1):
                if grp != span_name:
                    kpi_type_spans.append((span_name, span_start, i - 1))
                    span_name = grp
                    span_start = i
            kpi_type_spans.append((span_name, span_start, len(kpi_type_groups) - 1))

        # Make KPI names unique while preserving visible text by appending
        # zero-width-space suffixes to repeated labels. This prevents
        # dictionary-based row assembly from collapsing duplicate columns.
        seen = {}
        unique_kpi_names = []
        for name in kpi_names:
            cnt = seen.get(name, 0)
            if cnt == 0:
                unique_kpi_names.append(name)
            else:
                unique_kpi_names.append(name + '\u200b' * cnt)
            seen[name] = cnt + 1
        kpi_names = unique_kpi_names

        # Build set of column indices whose KPI header contains '%'.
        # openpyxl returns percentage-formatted cells as decimals (e.g. 0.85 for 85%);
        # multiply by 100 so values display as the user entered them in Excel.
        # A column is treated as percentage if:
        #   (a) its KPI header name contains '%', OR
        #   (b) any data cell in that column has a percentage Excel number format.
        # openpyxl stores percentage values as decimals (1.0 = 100%); we multiply by 100.
        pct_col_indices = set()
        for i, (col_idx, name) in enumerate(zip(included_cols, kpi_names)):
            if '%' in name:
                pct_col_indices.add(col_idx)
            else:
                # Scan first few data rows for a % number format on that column
                for scan_row in range(data_row_start, min(data_row_start + 5, ws.max_row + 1)):
                    cell_obj = merged_cell_obj(scan_row, col_idx)
                    if cell_obj is not None and cell_obj.value is not None and getattr(cell_obj, 'number_format', None) and '%' in cell_obj.number_format:
                        pct_col_indices.add(col_idx)
                        break

        # Get data from data_row_start to heatmap_max_row for the heatmap
        for row_idx in range(data_row_start, heatmap_max_row + 1):
            program_cell = merged_val(row_idx, program_col)
            if program_cell is not None and program_cell != "None":
                programs.append(str(program_cell))
                program_groups.append(b_values.get(row_idx) or "" if group_col is not None else "")
                row_data = []
                orig_data = []  # stores (fval, raw_val) tuples
                for col_idx in included_cols:
                    val = merged_val(row_idx, col_idx)
                    try:
                        fval = float(val) if val is not None else None
                        if fval is not None and col_idx in pct_col_indices:
                            fval = fval * 100
                            val = fval  # keep raw consistent
                    except:
                        fval = None
                    # Store NaN only for truly missing values (None); 0 is a valid value
                    row_data.append(fval if fval is not None else np.nan)
                    orig_data.append((fval, val))  # keep raw value to detect N/A text
                data_values.append(row_data)
                original_values.append(orig_data)

        # Collect rows beyond heatmap_max_row — merged into heatmap as gray rows
        if include_below_rows:
            for row_idx in range(heatmap_max_row + 1, ws.max_row + 1):
                program_cell = merged_val(row_idx, program_col)
                # Only include this below-row if it contains any non-blank data in the
                # KPI columns; otherwise skip so we don't render empty grey rows.
                # Treat None or empty/whitespace-only strings as blank.
                if program_cell is not None and str(program_cell).strip() not in ('', 'None'):
                    has_any = False
                    for col_idx in included_cols:
                        v = merged_val(row_idx, col_idx)
                        if v is None:
                            continue
                        if isinstance(v, str) and v.strip() == '':
                            continue
                        has_any = True
                        break
                    if not has_any:
                        continue
                    below_programs.append(str(program_cell))
                    row_data = []
                    orig_row = []  # stores (fval, raw_val) tuples
                    for col_idx in included_cols:
                        val = merged_val(row_idx, col_idx)
                        try:
                            fval = float(val) if val is not None else None
                            if fval is not None and col_idx in pct_col_indices:
                                fval = fval * 100
                                val = fval
                        except:
                            fval = None
                        row_data.append(fval if fval is not None else np.nan)
                        orig_row.append((fval, val))
                    below_data.append(row_data)
                    below_orig.append(orig_row)

        df_below = None  # no separate table

        # Merge all rows; track which are "below" for gray coloring
        all_programs = programs + below_programs
        all_data_values = data_values + below_data
        all_original_values = original_values + below_orig
        all_program_groups = program_groups + [''] * len(below_programs)

        # Compute program group spans for left-side bands
        program_group_spans = []
        if show_row_groups and all_program_groups:
            span_start = 0
            span_name = all_program_groups[0]
            for i, grp in enumerate(all_program_groups[1:], 1):
                if grp != span_name:
                    program_group_spans.append((span_name, span_start, i - 1))
                    span_name = grp
                    span_start = i
            program_group_spans.append((span_name, span_start, len(all_program_groups) - 1))

        # y-axis labels: plain program names only (group shown as side band)
        y_labels = all_programs

        # X-axis: word-wrapped bold labels at -45�
        MAX_CHARS = 15  # chars per line before wrapping
        CHAR_PX   = 7   # approx px per char at 10pt font
        def bold_wrap(text):
            lines = wrap_label(text, max_len=MAX_CHARS).split('<br>')
            return '<br>'.join(f'<b>{line}</b>' for line in lines)
        kpi_tick_names = [bold_wrap(k) for k in kpi_names]
        # Y-axis: word-wrapped bold labels (horizontal)
        y_labels_wrapped = ['<br>'.join(f'<b>{line}</b>' for line in wrap_label(p, max_len=24).split('<br>')) for p in all_programs]

        # Dynamically find the Per Program Target row by searching col 1 to program_col for the label
        per_program_target_row = None
        for row_idx in range(heatmap_max_row + 1, ws.max_row + 1):
            for col_idx in range(1, (program_col or 3) + 1):
                val = merged_val(row_idx, col_idx)
                if val and 'per program' in str(val).lower():
                    per_program_target_row = row_idx
                    break
            if per_program_target_row:
                break

        per_program_targets = []
        if per_program_target_row:
            for col_idx in range(data_col_start, data_col_start + len(kpi_names)):
                val = merged_val(per_program_target_row, col_idx)
                try:
                    per_program_targets.append(float(val) if val is not None else None)
                except:
                    per_program_targets.append(None)
        else:
            per_program_targets = [None] * len(kpi_names)

        # Fallback: if Per Program target cells are empty (e.g. uncached formulas),
        # compute from Annual Target row divided by number of heatmap programs
        if all(t is None for t in per_program_targets) and len(programs) > 0:
            annual_target_row = None
            for row_idx in range(heatmap_max_row + 1, ws.max_row + 1):
                for col_idx in range(1, (program_col or 3) + 1):
                    val = merged_val(row_idx, col_idx)
                    if val and 'annual target' in str(val).lower():
                        annual_target_row = row_idx
                        break
                if annual_target_row:
                    break
            if annual_target_row:
                n = len(programs)
                per_program_targets = []
                for col_idx in range(data_col_start, data_col_start + len(kpi_names)):
                    val = merged_val(annual_target_row, col_idx)
                    try:
                        annl = float(val) if val is not None else None
                        per_program_targets.append(annl / n if annl is not None and n > 0 else None)
                    except:
                        per_program_targets.append(None)
                # Patch below_orig so display row shows computed values not NA
                for bi, prog in enumerate(below_programs):
                    if 'per program' in prog.lower():
                        below_orig[bi] = [(t, t) for t in per_program_targets]
                        break
                # Re-merge after patch
                all_original_values = original_values + below_orig

        # Create dataframe using all rows
        if all_data_values and len(kpi_names) > 0:
            df_heatmap = pd.DataFrame(all_data_values, columns=kpi_names[:len(all_data_values[0])])
            df_heatmap.index = y_labels

            # Normalize using 3-point scale per column:
            #   min_val (lowest in rows 4-16) -> 0.0 (red)
            #   midpoint (per_program_target / 2)  -> 0.5 (yellow)
            #   per_program_target                 -> 1.0 (dark green)
            df_normalized = df_heatmap.copy()
            # Mark side columns (absolute Excel indices in side_cols) as NaN for data rows so they are not colored
            if side_cols:
                try:
                    for sc in side_cols:
                        if sc in included_cols:
                            pos = included_cols.index(sc)
                            if pos < df_normalized.shape[1]:
                                # only NaN for the main heatmap rows; below-rows will get -1 later
                                df_normalized.iloc[:len(programs), pos] = np.nan
                except Exception:
                    pass
            for col_i, col in enumerate(df_normalized.columns):
                # If this column maps to a declared side column, skip normalization (leave NaN)
                try:
                    if side_cols and col_i < len(included_cols) and included_cols[col_i] in side_cols:
                        df_normalized.iloc[:, col_i] = np.nan
                        continue
                except Exception:
                    pass
                # Use orig values (includes 0.0 correctly) to compute min
                hm_orig_floats = pd.Series(
                    [all_original_values[r][col_i][0] for r in range(len(programs))
                     if r < len(all_original_values) and col_i < len(all_original_values[r])
                     and isinstance(all_original_values[r][col_i], tuple)
                     and all_original_values[r][col_i][0] is not None],
                    dtype=float
                ).dropna()
                min_val  = hm_orig_floats.min() if len(hm_orig_floats) else 0.0
                target   = per_program_targets[col_i] if col_i < len(per_program_targets) and per_program_targets[col_i] else None
                midpoint = target / 2.0 if target else None
                for r in range(len(all_programs)):
                    if r >= len(programs):
                        df_normalized.iloc[r, col_i] = -1.0  # sentinel → grey
                        continue
                    orig_item = all_original_values[r][col_i] if r < len(all_original_values) and col_i < len(all_original_values[r]) else (None, None)
                    orig_fval = orig_item[0] if isinstance(orig_item, tuple) else None
                    orig_raw  = orig_item[1] if isinstance(orig_item, tuple) else orig_item
                    is_na_text = isinstance(orig_raw, str) and orig_raw.strip().upper() in ('N/A', 'NA', '#N/A')
                    is_na = is_na_text or orig_fval is None
                    if is_na:
                        df_normalized.iloc[r, col_i] = np.nan
                    elif target is None or target == min_val:
                        # No target defined — numeric values shown as green
                        df_normalized.iloc[r, col_i] = 1.0
                    elif orig_fval >= target:
                        # At or above target → always green (check before min_val)
                        df_normalized.iloc[r, col_i] = 1.0
                    elif orig_fval <= min_val:
                        df_normalized.iloc[r, col_i] = 0.0
                    elif midpoint and orig_fval <= midpoint:
                        span = midpoint - min_val
                        df_normalized.iloc[r, col_i] = 0.5 * (orig_fval - min_val) / span if span > 0 else 0.25
                    else:
                        span = target - midpoint
                        df_normalized.iloc[r, col_i] = 0.5 + 0.5 * (orig_fval - midpoint) / span if span > 0 else 0.75

            # Previously we forced entire below-row rows to grey. Remove that
            # behavior so that below-row cells containing actual data are
            # colored using the same normalization rules as main rows, while
            # blank/NA cells remain uncolored.

            # Extended colorscale: grey for below-row sentinel (-1), then red→green for data (0..1)
            # With zmin=-1, zmax=1 the normalised position = (v+1)/2:
            #   v=-1  → pos 0.00  (grey, below-row rows)
            #   v= 0  → pos 0.50  (red,  data min)
            #   v= 0.25 → pos 0.625 (orange)
            #   v= 0.5 → pos 0.75  (yellow)
            #   v= 0.75 → pos 0.875 (light green)
            #   v= 1  → pos 1.00  (dark green)
            colorscale = [
                [0.0,   '#9E9E9E'],  # grey (below-row sentinel)
                [0.499, '#9E9E9E'],  # grey end
                [0.5,   '#D73027'],  # dark red  (data min)
                [0.625, '#F46D43'],  # orange
                [0.75,  '#FFFF00'],  # yellow
                [0.875, '#A6D96A'],  # light green
                [1.0,   '#1A7A1A'],  # dark green
            ]

            # Smart numeric formatter — preserves significant figures for small values
            def fmt_val(v):
                """Format a number to at most 3 decimal places with thousand separators.
                Integers display without decimal point; very small values use 2 s.f. scientific."""
                if v == 0:
                    return '0'
                abs_v = abs(v)
                if abs_v < 0.0005:
                    # Too small for 3 dp — use 2 sig-fig scientific
                    return f"{v:.2e}"
                elif v == int(v):
                    return f"{int(v):,}"          # whole number with commas
                else:
                    formatted = f"{v:,.3f}".rstrip('0').rstrip('.')  # up to 3 dp, commas, strip trailing zeros
                    return formatted

            # Build text display
            text_display = []
            # Build set of column positions for fixed decimal formatting
            zero_dp_positions = set()
            if zero_decimal_cols:
                for ci, name in enumerate(kpi_names):
                    if any(sub.lower() in name.lower() for sub in zero_decimal_cols):
                        zero_dp_positions.add(ci)
            one_dp_positions = set()
            if one_decimal_cols:
                for ci, name in enumerate(kpi_names):
                    if any(sub.lower() in name.lower() for sub in one_decimal_cols):
                        one_dp_positions.add(ci)
            two_dp_positions = set()
            if two_decimal_cols:
                for ci, name in enumerate(kpi_names):
                    if any(sub.lower() in name.lower() for sub in two_decimal_cols):
                        two_dp_positions.add(ci)
            # Build set of row indices whose label matches zero_decimal_rows or one_decimal_rows substrings
            zero_dp_row_indices = set()
            if zero_decimal_rows:
                for ri, prog in enumerate(all_programs):
                    if any(sub.lower() in prog.lower() for sub in zero_decimal_rows):
                        zero_dp_row_indices.add(ri)
            one_dp_row_indices = set()
            if one_decimal_rows:
                for ri, prog in enumerate(all_programs):
                    if any(sub.lower() in prog.lower() for sub in one_decimal_rows):
                        one_dp_row_indices.add(ri)
            for row_i, row in enumerate(all_original_values):
                text_row = []
                row_force_zero_dp = row_i in zero_dp_row_indices
                row_force_one_dp = row_i in one_dp_row_indices
                for col_i, item in enumerate(row):
                    fval, raw = item if isinstance(item, tuple) else (item, None)
                    col_is_pct = col_i < len(included_cols) and included_cols[col_i] in pct_col_indices
                    # Show blank for truly empty cells; show 'NA' for N/A text; otherwise show raw text
                    is_na_text = isinstance(raw, str) and raw.strip().upper() in ('N/A', 'NA', '#N/A')
                    if fval is None:
                        # Preserve blank cells
                        if raw is None or (isinstance(raw, str) and raw.strip() == ''):
                            text_row.append('')
                        elif is_na_text:
                            text_row.append('NA')
                        else:
                            # non-numeric text that isn't N/A — show as-is
                            text_row.append(str(raw))
                    else:
                        if col_i in one_dp_positions:
                            formatted = f"{fval:,.1f}"
                        elif col_i in two_dp_positions:
                            formatted = f"{fval:,.2f}"
                        elif one_decimal_first_col and col_i == 0:
                            formatted = f"{fval:,.1f}"
                        elif row_force_one_dp:
                            formatted = f"{fval:,.1f}"
                        elif row_force_zero_dp or col_i in zero_dp_positions:
                            formatted = f"{round(fval):,}"
                        elif force_decimals is not None:
                            formatted = f"{fval:,.{force_decimals}f}"
                        else:
                            formatted = fmt_val(fval)
                        show_pct = col_is_pct and not suppress_pct_display
                        text_row.append(formatted + '%' if show_pct else formatted)
                text_display.append(text_row)

            # If requested, pad numeric text to align right using a monospace font.
            if monospace_numeric:
                ncols = len(kpi_names)
                nrows = len(text_display)
                col_max = [0] * ncols
                for j in range(ncols):
                    for i in range(nrows):
                        if j < len(text_display[i]):
                            s = str(text_display[i][j])
                        else:
                            s = ''
                        if len(s) > col_max[j]:
                            col_max[j] = len(s)
                for i in range(nrows):
                    for j in range(ncols):
                        if j < len(text_display[i]):
                            s = str(text_display[i][j])
                        else:
                            s = ''
                        text_display[i][j] = s.rjust(col_max[j], '\u00A0')
            text_family = 'Courier New, monospace' if monospace_numeric else 'Arial, sans-serif'

            fig = px.imshow(
                df_normalized.values,
                x=kpi_names,
                y=y_labels,
                labels=dict(x="KPI", y="Program", color="Value"),
                color_continuous_scale=colorscale,
                text_auto=False,
                aspect="auto",
                zmin=-1,
                zmax=1
            )

            # Overlay text (NA or rounded value) and add cell grid via gaps
            fig.update_traces(
                text=np.array(text_display, dtype=object),
                texttemplate='%{text}',
                textfont=dict(size=16, color='black', family=text_family),
                xgap=2,
                ygap=2
            )

            # Add a separate text-only trace for below-row cells so their text is always visible on grey
            if len(below_programs) > 0:
                below_text = text_display[len(programs):]
                fig.add_trace(go.Heatmap(
                    z=np.full((len(below_programs), len(kpi_names)), np.nan),
                    x=kpi_names,
                    y=all_programs[len(programs):],
                    text=np.array(below_text, dtype=object),
                    texttemplate='%{text}',
                    textfont=dict(size=16, color='#333333', family=text_family),
                    showscale=False,
                    coloraxis=None,
                    colorscale=[[0, 'rgba(0,0,0,0)'], [1, 'rgba(0,0,0,0)']],
                    zmin=0, zmax=1,
                    xgap=2, ygap=2,
                    hoverinfo='skip',
                ))

            # LABEL_PX: at -45� with wrapped labels, project max lines * line_height * sin(45�)
            LINE_H_PX = 12  # approx line height in px at 9pt
            max_lines = max((len(t.split('<br>')) for t in kpi_tick_names), default=1)
            longest_line = max((len(line.replace('<b>','').replace('</b>','')) for t in kpi_tick_names for line in t.split('<br>')), default=10)
            # Auto-detect short labels (e.g. years) vs long KPI names
            short_labels = longest_line <= 6 and max_lines == 1
            if short_labels:
                LABEL_PX = 30
                tick_angle = 0
                col_px = 35
            else:
                LABEL_PX = max(60, int((longest_line * CHAR_PX + max_lines * LINE_H_PX) * 0.71) + 10)
                tick_angle = -45
                col_px = 80  # wide enough so diagonal labels don't override each other
            BAND_PX  = 30
            GAP_PX   = group_gap if group_gap is not None else (8 if short_labels else 0)  # extra space between tick labels and group bands
            dynamic_top = LABEL_PX + GAP_PX + BAND_PX + (extra_top or 0)
            row_px = 32  # taller rows so text is more readable
            # Compute left margin based on longest program label so row text fits
            import re
            clean_y_labels = [re.sub(r'<[^>]+>', '', t) for t in y_labels_wrapped]
            longest_label_chars = max((len(s) for s in clean_y_labels), default=18)
            # increase base padding to give more horizontal room for long labels
            computed_left = int(longest_label_chars * CHAR_PX + 160)
            # bump defaults slightly
            LEFT_M = left_margin if left_margin is not None else max(computed_left, 400 if show_row_groups else 260)
            RIGHT_M = 20
            BOTTOM_M = 20
            chart_height = dynamic_top + 40 + len(all_programs) * row_px
            chart_width  = LEFT_M + RIGHT_M + len(kpi_names) * col_px

            # Paper coords above y=1.0 use the PLOT AREA height (not total chart height)
            # plot_area_height = chart_height - top_margin - bottom_margin
            plot_area_h = chart_height - dynamic_top - BOTTOM_M
            band_y0 = 1.0 + (LABEL_PX + GAP_PX) / plot_area_h
            band_y1 = band_y0 + BAND_PX / plot_area_h
            band_label_y = (band_y0 + band_y1) / 2

            # Move x-axis to top with wrapped tick labels
            fig.update_layout(
                height=chart_height,
                width=chart_width,
                xaxis=dict(
                    type='category',
                    side='top',
                    tickangle=tick_angle,
                    title='',
                    tickmode='array',
                    tickvals=kpi_names,
                    ticktext=kpi_tick_names,
                    tickfont=dict(size=10 if short_labels else 9),
                    automargin=False,
                    showgrid=False,
                    zeroline=False,
                    showline=False,
                ),
                yaxis=dict(
                    tickfont=dict(size=9),
                    tickmode='array',
                    tickvals=all_programs,
                    ticktext=y_labels_wrapped,
                    automargin=True,
                    showgrid=False,
                    zeroline=False,
                    showline=False,
                ),
                yaxis_title="",
                title="",
                margin=dict(l=LEFT_M, r=RIGHT_M, t=dynamic_top, b=BOTTOM_M),
                coloraxis_showscale=False
            )

            # Add KPI type group header rectangles + labels above the x-axis
            # Positions are dynamic (band_y0/y1/label_y) to avoid overlapping tick labels
            top_group_colors = [
                '#007a17', '#00891a', '#005c11', '#006b14',
                '#004f0e', '#008a1a', '#003d0b', '#009e1e'
            ]
            left_group_colors = [
                '#005a8e', '#004470', '#2471a3', '#1a5f8a',
                '#003d6b', '#1a6fa8', '#002e52', '#0d5496'
            ]
            for idx, (group_name, start_idx, end_idx) in enumerate(kpi_type_spans):
                color = top_group_colors[idx % len(top_group_colors)]
                x_center = (start_idx + end_idx) / 2
                # Top group: no fill, larger black label
                fig.add_shape(
                    type='rect',
                    xref='x', yref='paper',
                    x0=start_idx - 0.5, x1=end_idx + 0.5,
                    y0=band_y0, y1=band_y1,
                    fillcolor='rgba(0,0,0,0)',
                    line=dict(color='rgba(0,0,0,0)', width=0),
                    layer='above'
                )
                # Wrap group name so long labels don't overflow their column span
                wrapped_group = '<br>'.join(f'<b>{line}</b>' for line in wrap_label(group_name, max_len=20).split('<br>'))
                fig.add_annotation(
                    xref='x', yref='paper',
                    x=x_center, y=band_label_y,
                    text=wrapped_group,
                    showarrow=False,
                    font=dict(color='black', size=14, family='Arial Black, Arial, sans-serif'),
                    align='center',
                    xanchor='center',
                    yanchor='middle',
                    bgcolor='rgba(0,0,0,0)'
                )
                # Add a thick black vertical line at the end of this column group
                # Extend from bottom of heatmap (y=1.0 paper) up through the column group band (band_y1)
                fig.add_shape(
                    type='line',
                    xref='x', yref='paper',
                    x0=end_idx + 0.5, x1=end_idx + 0.5,
                    y0=0.0, y1=band_y1,
                    line=dict(color='black', width=2),
                    layer='above'
                )

            # Add program group bands to the LEFT of the y-axis (mirrors top KPI-type bands)
            if show_row_groups:
                # Compute band x positions dynamically so they scale correctly with LEFT_M
                # and the plot area width — prevents crowding when plot area is narrow.
                _plot_w_px = len(kpi_names) * col_px
                # Move group band further into the left margin to create a visible gap
                gx_outer = -(LEFT_M * 0.75) / _plot_w_px   # outer edge (furthest into margin)
                gx_inner = -(LEFT_M * 0.25) / _plot_w_px   # inner edge (closest to y-axis)
                gx_ann   = -(LEFT_M * 0.60) / _plot_w_px  # annotation midpoint (further left)
                for idx, (group_name, start_idx, end_idx) in enumerate(program_group_spans):
                    # Skip bands that are entirely beyond the main heatmap rows or have no group name
                    if start_idx >= len(programs) or not group_name:
                        continue
                    end_idx = min(end_idx, len(programs) - 1)
                    color = left_group_colors[idx % len(left_group_colors)]
                    y_center = (start_idx + end_idx) / 2
                    # Left group: no fill, larger black horizontal label anchored to the right
                    fig.add_shape(
                        type='rect',
                        xref='paper', yref='y',
                        x0=gx_outer, x1=gx_inner,
                        y0=start_idx - 0.5, y1=end_idx + 0.5,
                        fillcolor='rgba(0,0,0,0)',
                        line=dict(color='rgba(0,0,0,0)', width=0),
                        layer='above'
                    )
                    # Wrap long group names so they don't overflow and anchor to the right
                    wrapped_left = '<br>'.join(f'<b>{line}</b>' for line in wrap_label(group_name, max_len=12).split('<br>'))
                    fig.add_annotation(
                        xref='paper', yref='y',
                        x=gx_ann, y=y_center,
                        text=wrapped_left,
                        showarrow=False,
                        font=dict(color='black', size=12, family='Arial Black, Arial, sans-serif'),
                        align='center',
                        xanchor='right',
                        textangle=0,
                        bgcolor='rgba(0,0,0,0)'
                    )
                    # Add a thick black horizontal line at the end of this row group
                    # Extend from outer edge of row-group band to right edge of heatmap
                    fig.add_shape(
                        type='line',
                        xref='paper', yref='y',
                        x0=gx_outer, x1=1.0,
                        y0=end_idx + 0.5, y1=end_idx + 0.5,
                        line=dict(color='black', width=2),
                        layer='above'
                    )

            # Build raw verification DataFrame (what was read from Excel)
            raw_display_rows = []
            for r, prog in enumerate(all_programs):
                row_dict = {'Program': prog}
                for col_i, kpi in enumerate(kpi_names):
                    if r < len(text_display) and col_i < len(text_display[r]):
                        row_dict[kpi] = text_display[r][col_i]
                    else:
                        row_dict[kpi] = 'NA'
                raw_display_rows.append(row_dict)
            df_raw = pd.DataFrame(raw_display_rows).set_index('Program') if raw_display_rows else None

            return fig, df_below, df_raw
        else:
            st.error(f"No data found. Programs: {len(programs)}, KPIs: {len(kpi_names)}")
            return None, None, None
    except Exception as e:
        st.error(f"Error creating heatmap: {str(e)}")
        return None, None, None

# Helper: render a dataframe as a gray-styled HTML table
def render_gray_table(df):
    header_cells = "".join(
        f'<th style="background-color:#6b7280;color:white;padding:8px 12px;border:1px solid #9ca3af;font-weight:bold;">{col}</th>'
        for col in df.columns
    )
    rows_html = ""
    for i, row in df.iterrows():
        bg = "#f3f4f6" if i % 2 == 0 else "#e5e7eb"
        cells = "".join(
            f'<td style="background-color:{bg};padding:7px 12px;border:1px solid #d1d5db;">{val}</td>'
            for val in row
        )
        rows_html += f"<tr>{cells}</tr>"
    html = f"""
    <div style="overflow-x:auto;">
    <table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:13px;">
      <thead><tr>{header_cells}</tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    </div>"""
    st.markdown(html, unsafe_allow_html=True)

# Load data
df_programs, df_services, df_heatmap = load_kpi_data()

# Tabs
tab1, tab2, tab3 = st.tabs(["📊 2025 Program Output KPIs (Aggregate)", "🌡️ 2025 Program Output KPI (by Program)", "🏢 2025 Service Unit KPIs"])

# Programs Tab
with tab1:
    st.subheader("📊 2025 Program Output KPIs (Aggregate)")
    
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    program_file = os.path.join(root_dir, 'data', 'Program Output KPIs.xlsx')
    
    try:
        html_programs = excel_to_html_with_merged_cells(program_file, no_decimals=True)
        st.markdown(html_programs, unsafe_allow_html=True)
    except Exception as e:
        st.warning(f"Could not render with merged cells: {str(e)}")
        # Fallback: format numeric columns to have no decimals and convert to strings
        display_df = df_programs.copy()
        for col in display_df.select_dtypes(include=["number"]).columns:
            display_df[col] = display_df[col].apply(lambda x: "" if pd.isna(x) else str(int(round(x))))
        st.dataframe(display_df, width='stretch', height=600)
    
    # Download button
    csv_programs = df_programs.to_csv(index=False)
    st.download_button(
        label="⬇️ Download 2025 Program KPIs as CSV",
        data=csv_programs,
        file_name="2025_Program_Output_KPIs.csv",
        mime="text/csv"
    )

# KPI By Program Tab (now second)
with tab2:
    st.subheader("🌡️ 2025 Program Output KPI (by Program)")
    
    # Two main sub-tabs
    sub_tab_a, sub_tab_b = st.tabs(["🔬 Research, Training, Product Development", "🏆 Recognition, Societal Impact & Inclusivity"])
    
    # ==================== Research, Training, Product Development ====================
    with sub_tab_a:
        st.markdown("### Research, Training, Product Development")
        
        rtpd_tabs = st.tabs(["KPI by Nr", "KPI by FTE", "KPI by $", "KPI over Time"])
        
        root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        with rtpd_tabs[0]:
            st.write("**Research, Training, Product Development - KPI by Nr**")
            try:
                heatmap_file = os.path.join(root_dir, 'data', 'Heat map 1.xlsx')
                if os.path.exists(heatmap_file):
                    fig, df_below, df_raw = create_heatmap_visualization(heatmap_file, zero_decimal_cols=['Thompson', 'Thomson'], one_decimal_cols=['per IRS', 'per irs'], zero_decimal_rows=['per program target', 'per programme target'])
                    if fig:
                        st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                    if df_below is not None and not df_below.empty:
                        st.markdown("---")
                        st.markdown("**Additional Data**")
                        render_gray_table(df_below)
                else:
                    st.info("📁 Waiting for: Heat map 1.xlsx")
            except Exception as e:
                st.warning(f"Could not load heatmap: {str(e)}")

        with rtpd_tabs[1]:
            st.write("**Research, Training, Product Development - KPI by FTE**")
            try:
                heatmap_file = os.path.join(root_dir, 'data', 'Heat map 2.xlsx')
                if os.path.exists(heatmap_file):
                    fig, df_below, df_raw = create_heatmap_visualization(heatmap_file, side_cols=[4], force_decimals=2, suppress_pct_display=True, one_decimal_first_col=True, no_gray_first_col=True)
                    if fig:
                        st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                    if df_below is not None and not df_below.empty:
                        st.markdown("---")
                        st.markdown("**Additional Data**")
                        render_gray_table(df_below)
                else:
                    st.info("📁 Waiting for: Heat map 2.xlsx")
            except Exception as e:
                st.warning(f"Could not load heatmap: {str(e)}")

        with rtpd_tabs[2]:
            st.write("**Research, Training, Product Development - KPI by $**")
            try:
                heatmap_file = os.path.join(root_dir, 'data', 'Heat map 3.xlsx')
                if os.path.exists(heatmap_file):
                    fig, df_below, df_raw = create_heatmap_visualization(heatmap_file, side_cols=[4], force_decimals=2, monospace_numeric=False, one_decimal_first_col=True, no_gray_first_col=True)
                    if fig:
                        st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                    if df_below is not None and not df_below.empty:
                        st.markdown("---")
                        st.markdown("**Additional Data**")
                        render_gray_table(df_below)
                else:
                    st.info("📁 Waiting for: Heat map 3.xlsx")
            except Exception as e:
                st.warning(f"Could not load heatmap: {str(e)}")

        with rtpd_tabs[3]:
            st.write("**Research, Training, Product Development - KPI over Time**")
            try:
                heatmap_file = os.path.join(root_dir, 'data', 'Heat map 4.xlsx')
                heatmap_file_4_1 = os.path.join(root_dir, 'data', 'Heat map 4-1 Research Outputs.xlsx')
                heatmap_choice = heatmap_file_4_1 if os.path.exists(heatmap_file_4_1) else heatmap_file
                if os.path.exists(heatmap_choice):
                    st.caption(f"Using file: {os.path.basename(heatmap_choice)}")
                    # Provide three focused KPI-over-time sub-tabs so users can
                    # view Research Outputs, Capacity Building, and Product
                    # Development separately.
                    ot_tabs = st.tabs(["Research Outputs", "Capacity Building", "Product Development"])

                    with ot_tabs[0]:
                        # Research Outputs
                        fig, df_below, df_raw = create_heatmap_visualization(
                            heatmap_choice,
                            zero_decimal_cols=['Thompson', 'Thomson'],
                            one_decimal_cols=['per IRS', 'per irs'],
                            zero_decimal_rows=['per program target', 'per programme target'],
                            force_decimals=0,
                            kpi_group_filter='research',
                            kpi_group_row=2
                        )
                        if fig:
                            st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                        if df_below is not None and not df_below.empty:
                            st.markdown("---")
                            st.markdown("**Additional Data**")
                            render_gray_table(df_below)

                    with ot_tabs[1]:
                        # Capacity Building — prefer a dedicated capacity-building file if present
                        cap_file = os.path.join(root_dir, 'data', 'Heat map 4-1 Capacity Building.xlsx')
                        cap_choice = cap_file if os.path.exists(cap_file) else heatmap_choice
                        # Data is known to live in columns 10,16,22 for this sheet
                        fig, df_below, df_raw = create_heatmap_visualization(
                            cap_choice,
                            zero_decimal_cols=['Thompson', 'Thomson'],
                            one_decimal_cols=['per IRS', 'per irs'],
                            zero_decimal_rows=['per program target', 'per programme target'],
                            force_decimals=0,
                            kpi_group_filter='capacity',
                            force_include_cols=[10, 16, 22],
                            kpi_row=3,
                            kpi_group_row=2
                        )
                        if fig:
                            st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                        if df_below is not None and not df_below.empty:
                            st.markdown("---")
                            st.markdown("**Additional Data**")
                            render_gray_table(df_below)

                    with ot_tabs[2]:
                        # Product Development — prefer a dedicated product-development file if present
                        prod_file = os.path.join(root_dir, 'data', 'Heat map 4-1 - Product Development.xlsx')
                        prod_choice = prod_file if os.path.exists(prod_file) else heatmap_choice
                        # Data is known to live in columns 10,16,22,28 for this sheet
                        fig, df_below, df_raw = create_heatmap_visualization(
                            prod_choice,
                            zero_decimal_cols=['Thompson', 'Thomson'],
                            one_decimal_cols=['per IRS', 'per irs'],
                            zero_decimal_rows=['per program target', 'per programme target'],
                            force_decimals=0,
                            kpi_group_filter='product',
                            force_include_cols=[10, 16, 22, 28],
                            kpi_row=3,
                            kpi_group_row=2,
                            extra_top=30,
                            group_gap=40
                        )
                        if fig:
                            st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                        if df_below is not None and not df_below.empty:
                            st.markdown("---")
                            st.markdown("**Additional Data**")
                            render_gray_table(df_below)
                else:
                    st.info("📁 Waiting for: Heat map 4.xlsx")
            except Exception as e:
                st.warning(f"Could not load heatmap: {str(e)}")
    
    # ==================== Recognition, Societal Impact & Inclusivity ====================
    with sub_tab_b:
        st.markdown("### Recognition, Societal Impact & Inclusivity")
        
        rsi_tabs = st.tabs(["KPI by Nr", "KPI by FTE", "KPI by $", "KPI over Time"])
        
        root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        with rsi_tabs[0]:
            st.write("**Recognition, Societal Impact & Inclusivity - KPI by Nr**")
            try:
                heatmap_file = os.path.join(root_dir, 'data', 'Heat map 5.xlsx')
                if os.path.exists(heatmap_file):
                    fig, df_below, df_raw = create_heatmap_visualization(heatmap_file, left_margin=400, one_decimal_rows=['per program target', 'per programme target'])
                    if fig:
                        st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                    if df_below is not None and not df_below.empty:
                        st.markdown("---")
                        st.markdown("**Additional Data**")
                        render_gray_table(df_below)
                else:
                    st.info("📁 Waiting for: Heat map 5.xlsx")
            except Exception as e:
                st.warning(f"Could not load heatmap: {str(e)}")

        with rsi_tabs[1]:
            st.write("**Recognition, Societal Impact & Inclusivity - KPI by FTE**")
            try:
                heatmap_file = os.path.join(root_dir, 'data', 'Heat map 6.xlsx')
                if os.path.exists(heatmap_file):
                    fig, df_below, df_raw = create_heatmap_visualization(heatmap_file, side_cols=[4], left_margin=560, one_decimal_first_col=True, force_decimals=3, no_gray_first_col=True)
                    if fig:
                        st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                    if df_below is not None and not df_below.empty:
                        st.markdown("---")
                        st.markdown("**Additional Data**")
                        render_gray_table(df_below)
                else:
                    st.info("📁 Waiting for: Heat map 6.xlsx")
            except Exception as e:
                st.warning(f"Could not load heatmap: {str(e)}")

        with rsi_tabs[2]:
            st.write("**Recognition, Societal Impact & Inclusivity - KPI by $**")
            try:
                heatmap_file = os.path.join(root_dir, 'data', 'Heat map 7.xlsx')
                if os.path.exists(heatmap_file):
                    fig, df_below, df_raw = create_heatmap_visualization(heatmap_file, side_cols=[4], left_margin=560, one_decimal_first_col=True, force_decimals=3, no_gray_first_col=True)
                    if fig:
                        st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                    if df_below is not None and not df_below.empty:
                        st.markdown("---")
                        st.markdown("**Additional Data**")
                        render_gray_table(df_below)
                else:
                    st.info("📁 Waiting for: Heat map 7.xlsx")
            except Exception as e:
                st.warning(f"Could not load heatmap: {str(e)}")

        with rsi_tabs[3]:
            st.write("**Recognition, Societal Impact & Inclusivity - KPI over Time**")
            # Create focused sub-tabs for Recognition and Societal Impact
            rsi_ot_sub = st.tabs(["Recognition and Reputation", "Societal Impact and Inclusion"]) 
            # Base heatmap file fallback
            heatmap_file = os.path.join(root_dir, 'data', 'Heat map 5.xlsx')

            with rsi_ot_sub[0]:
                try:
                    # Prefer dedicated Heat map 4-2 recognition file; fall back to Heat map 5
                    rec_file = os.path.join(root_dir, 'data', 'Heat map 4-2 Recognition and Reputation.xlsx')
                    rec_choice = rec_file if os.path.exists(rec_file) else heatmap_file
                    if os.path.exists(rec_choice):
                        fig, df_below, df_raw = create_heatmap_visualization(
                            rec_choice,
                            zero_decimal_cols=['Thompson', 'Thomson'],
                            one_decimal_cols=['per IRS', 'per irs'],
                            zero_decimal_rows=['per program target', 'per programme target'],
                            force_decimals=0,
                            kpi_group_filter='recognition',
                            kpi_group_row=2,
                            extra_top=30,
                            group_gap=40
                        )
                        if fig:
                            st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                        if df_below is not None and not df_below.empty:
                            st.markdown('---')
                            st.markdown('**Additional Data**')
                            render_gray_table(df_below)
                    else:
                        st.info('📁 Waiting for: Heat map 5.xlsx or Recognition file')
                except Exception as e:
                    st.warning(f"Could not load Recognition heatmap: {str(e)}")

            with rsi_ot_sub[1]:
                try:
                    soc_file = os.path.join(root_dir, 'data', 'Heat map 4-2 Society Impact and Inclusion.xlsx')
                    soc_choice = soc_file if os.path.exists(soc_file) else heatmap_file
                    if os.path.exists(soc_choice):
                        fig, df_below, df_raw = create_heatmap_visualization(
                            soc_choice,
                            zero_decimal_cols=['Thompson', 'Thomson'],
                            one_decimal_cols=['per IRS', 'per irs'],
                            zero_decimal_rows=['per program target', 'per programme target'],
                            force_decimals=0,
                            kpi_group_filter='societal',
                            kpi_group_row=2,
                            extra_top=30,
                            group_gap=40
                        )
                        if fig:
                            st.plotly_chart(fig, use_container_width=False, config={'scrollZoom': False})
                        if df_below is not None and not df_below.empty:
                            st.markdown('---')
                            st.markdown('**Additional Data**')
                            render_gray_table(df_below)
                    else:
                        st.info('📁 Waiting for: Heat map 5.xlsx or Societal Impact file')
                except Exception as e:
                    st.warning(f"Could not load Societal Impact heatmap: {str(e)}")

# Service Units Tab (now third)
with tab3:
    st.subheader("🏢 2025 Service Unit KPIs")
    
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    service_file = os.path.join(root_dir, 'data', 'Service Unit KPIs.xlsx')
    
    try:
        html_services = excel_to_html_with_merged_cells(service_file, no_decimals=True, highlight_row_keyword='service unit key performance')
        st.markdown(html_services, unsafe_allow_html=True)
    except Exception as e:
        st.warning(f"Could not render with merged cells: {str(e)}")
        # Fallback: format numeric columns to have no decimals and convert to strings
        display_df_s = df_services.copy()
        for col in display_df_s.select_dtypes(include=["number"]).columns:
            display_df_s[col] = display_df_s[col].apply(lambda x: "" if pd.isna(x) else str(int(round(x))))
        st.dataframe(display_df_s, width='stretch', height=600)
    # Download button
    csv_services = df_services.to_csv(index=False)
    st.download_button(
        label="⬇️ Download 2025 Service Unit KPIs as CSV",
        data=csv_services,
        file_name="2025_Service_Unit_KPIs.csv",
        mime="text/csv"
    )

st.markdown("---")
st.caption("Last updated: April 8, 2026 | IITA KPI Dashboard")
